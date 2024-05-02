import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU

from driver_work_report_generator.api import (TabelaParadaEspera, TabelaViagens, TabelaJornadaDiaria, TabelaMensagens,
                                              TabelaDescansoIntrajornada, TabelaHorarioRefeicao, TabelaCargaDecarga)
from driver_work_report_generator.settings import OUTPUT_REPORT
from driver_work_report_generator.utils import convert_to_time


FORMAT_TIME = '[HH]:MM:SS'
BLACK = "000000"
GREY = 'ffededed'
ALIGN = Alignment(horizontal='center', vertical='center')
FILL = PatternFill(start_color=GREY, end_color=GREY, fill_type='solid')


def save(data_dict, driver, filename, logo_image):
    dates = get_dates(data_dict, driver, 7)

    wb = openpyxl.load_workbook(filename)
    ws = wb[wb.sheetnames[0]]
    start_row = row = 9

    medium_border = Side(border_style="medium", color=BLACK)  # borda média preta

    for date, day in dates:
        data = f'{date} -\n{day.capitalize()}'
        diaria = get_value(date, TabelaJornadaDiaria, data_dict, driver, 6, 5)
        refeicao = get_value(date, TabelaHorarioRefeicao, data_dict, driver, 6, 5)
        carga_descarga = get_value(date, TabelaCargaDecarga, data_dict, driver, 6, 5)
        t_espera = get_value(date, TabelaParadaEspera, data_dict, driver, 6, 5)
        horas_direcao = get_value(date, TabelaViagens, data_dict, driver, 3, 4)
        hora_extra = get_value(date, TabelaViagens, data_dict, driver, 3, 5)
        if hora_extra and hora_extra.startswith('-'):
            hora_extra = '0:00:00'
        hora_noturna = get_value(date, TabelaJornadaDiaria, data_dict, driver, 6, 5, filter_plate=False, filter_shift=True)
        hora_interjornada = get_value(date, TabelaDescansoIntrajornada, data_dict, driver, 6, 5)

        ws.cell(row=row, column=1).value = data   # data
        ws.cell(row=row, column=2).value = convert_to_time(diaria)   # diária
        ws.cell(row=row, column=3).value = convert_to_time(refeicao)   # refeicao
        ws.cell(row=row, column=4).value = convert_to_time(carga_descarga)   # carga_descarga
        ws.cell(row=row, column=5).value = convert_to_time(t_espera)   # t_espera
        ws.cell(row=row, column=6).value = convert_to_time(horas_direcao)   # horas_direcao
        # ws.cell(row=row, column=7).value = convert_to_time(hora_extra)   # hora_extra
        # hora_extra
        ws.cell(row=row, column=7).value = f'=IF(ISNUMBER(SEARCH("domingo",A{row})),B{row},IF(ISNUMBER(SEARCH("sábado",A{row})),IF(B{row}<=TIME(4,0,0),TIME(0,0,0),B{row}-TIME(4,0,0)),IF(B{row}<=TIME(8,0,0),TIME(0,0,0),B{row}-TIME(8,0,0))))'
        # hora extra 60%
        ws.cell(row=row, column=8).value = f'=IF(ISNUMBER(SEARCH("domingo", A{row})), IF(B{row}>=TIME(2,0,0), TIME(2,0,0), B{row}), IF(ISNUMBER(SEARCH("sábado", A{row})), IF(B{row}<=TIME(4,0,0), TIME(0,0,0), IF(B{row}>=TIME(6,0,0), TIME(2,0,0), B{row}-TIME(4,0,0))),IF(B{row}<=TIME(8,0,0), TIME(0,0,0), IF(B{row}>=TIME(10,0,0), TIME(2,0,0), B{row}-TIME(8,0,0)))))'
        # hora extra 100%
        ws.cell(row=row, column=9).value = f'=IF(ISNUMBER(SEARCH("domingo",A{row})),IF(B{row}<=TIME(2,0,0),TIME(0,0,0), B{row} - TIME(2,0,0)), IF(ISNUMBER(SEARCH("SÁBADO",A{row})),IF(B{row}<=TIME(6,0,0),TIME(0,0,0), B{row} - TIME(6,0,0)), IF(B{row}<=TIME(10,0,0),TIME(0,0,0), B{row} - TIME(10,0,0))))'
        ws.cell(row=row, column=10).value = convert_to_time(hora_interjornada)   # hora_interjornada
        ws.cell(row=row, column=11).value = f'=IF(J{row}>=TIME(11,0,0), TIME(0,0,0) , TIME(11,0,0)-J{row})'   #  hora_interjornada_indenizada
        ws.cell(row=row, column=12).value = convert_to_time(hora_noturna)   # hora_noturna

        format_cells(ws, row)

        row += 1

    # TOTAIS
    ws.cell(row=row, column=1).value = 'TOTAL DE \nHORAS'
    ws.cell(row=row, column=1).border = Border(right=medium_border, bottom=medium_border, left=medium_border, top=medium_border)
    ws.cell(row=row, column=1).alignment = ALIGN
    ws.cell(row=row, column=1).fill = FILL
    ws.cell(row=row, column=1).font = Font('Arial', size=10, bold=True)
    for col_index in range(2, 13):
        col_letter = get_column_letter(col_index)
        cell = ws.cell(row=row, column=col_index)
        cell.number_format = FORMAT_TIME
        cell.value = f'=subtotal(9, {col_letter}{start_row}:{col_letter}{row-1})'
        cell.border = Border(right=medium_border, bottom=medium_border, left=medium_border, top=medium_border)
        cell.alignment = ALIGN
        cell.fill = FILL

    ws.cell(row=5, column=1).value = f'MOTORISTA:  {driver}'
    ws.row_dimensions[row].height = 30

    add_image(ws, logo_image)

    output_file = OUTPUT_REPORT

    wb.save(output_file)

    wb.close()

    return output_file


def prepare_data(class_tabela, data_dict, driver, pos, filter_plate=True, filter_shift=False):
    tabela = class_tabela()
    data = tabela.execute(data=data_dict)
    data = tabela._insert_date(data)
    data = tabela.filter_driver(data, driver, pos)

    if filter_plate:
        data = tabela.filter_plate_itens(data)
    if filter_shift:
        data = tabela.filter_night_shift_itens(data)

    return data


def get_value(date, tabela_class, data_dict, driver, driver_index, value_pos, filter_plate=True, filter_shift=False):
    value = None
    for item in prepare_data(tabela_class, data_dict, driver, driver_index, filter_plate, filter_shift):
        if date == item[0]:
            value = item[value_pos]
            break
    return value


def get_dates(data_dict, driver, driver_index):
    tabela = TabelaMensagens()
    data = tabela.execute(data=data_dict)
    data = tabela._insert_date(data)
    data = tabela.filter_plate_itens(data)
    data = tabela.filter_driver(data, driver, driver_index)
    return tabela.get_dates(data)


def format_cells(ws, row):
    for c in range(2, 13):
        ws.cell(row=row, column=c).number_format = FORMAT_TIME

    thin_border = Side(border_style="thin", color=BLACK)  # borda fina preta
    border = Border(right=thin_border, bottom=thin_border, left=thin_border, top=thin_border)
    for c in range(1, 13):
        cell = ws.cell(row=row, column=c)
        cell.border = border
        cell.alignment = ALIGN
        cell.fill = FILL

    ws.row_dimensions[row].height = 30


def add_image(ws, logo_image):
    p2e = pixels_to_EMU
    img = Image(logo_image)
    h, w = img.height, img.width

    position = XDRPoint2D(p2e(4), p2e(10))
    size = XDRPositiveSize2D(p2e(w), p2e(h))

    img.anchor = AbsoluteAnchor(pos=position, ext=size)
    ws.add_image(img)
